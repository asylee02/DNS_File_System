import pandas as pd
from openpyxl import load_workbook

file_named = 'DNS_Test.xlsx' # 파일 경로를 적절히 변경하세요

wb = load_workbook(filename=file_named)


# 숨긴 파일 & 추가 파일 삭제
def eliminate(wb):
    hidden_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'hidden']
    plus_sheets =[]
    for sheet_name in wb.sheetnames:
        if "추가" in sheet_name:
            inx=wb.sheetnames.index(sheet_name);
            plus_sheets.append(wb.sheetnames[inx-1])

    for sheet_name in hidden_sheets:
        del wb[sheet_name]
    for sheet_name in plus_sheets:
        del wb[sheet_name]
    del wb['SHIPPING MARK']
    wb.save('Test2.xlsx')

eliminate(wb)
#숨긴 시트 삭제된 파일 불러오기
file_path='Test2.xlsx'

excel_data = pd.ExcelFile(file_path)
pd.read_excel(file_path,  engine='openpyxl' )
sheet_List = list(excel_data.sheet_names)
# 시트의 개수와 이름 확인
num_sheets = len(sheet_List)
print(f"시트의 개수: {num_sheets}")
print(sheet_List)
total_cnt = 0

Total_name_List =[]
    

def remove_prefix(item):
    if isinstance(item[0], str):
        return [item[0].split('. ', 1)[1]] + item[1:]
    else:
        return item

def duplication(file):
    result_dict={}
    for item in file:
        if item[0] in result_dict:
            result_dict[item[0]] += item[1]
        else:
            result_dict[item[0]] = item[1]
    new_list = [[key, value] for key, value in result_dict.items()]
    return new_list

def Extraction(i, List):
    globals()[f'List{i}'] = []
    df1 = pd.read_excel(file_path, sheet_name=i, engine='openpyxl', usecols=List )
    cleaned_df = df1.dropna(axis=0, how='any')
    for index, row in cleaned_df.iterrows():
        row_values = row.values.tolist()
        globals()[f'List{i}'].append(row_values)    


# 단가
def Price(i):
    Extraction(i,[1,2,11])
    
# PI
def Proforma(i):
    Extraction(i,[2,6,8])
    globals()[f'List{i}'].pop(0)

# 발주서
def Order(i):
    Extraction(i,[2,6])
    Des_idx=[]
    for k in range(len(globals()[f'List{i}'])):
        if globals()[f'List{i}'][k][0] == 'Description':
            Des_idx.append(k)
    Des_idx.reverse()
    for j in Des_idx:
        globals()[f'List{i}'].pop(j)
 
# CI
def Commercial(i):
    Extraction(i,[1,5,7])
    delete_list = []
    for k in range(len(globals()[f'List{i}'])):
        if globals()[f'List{i}'][k][0] == 'Description of Goods':
            delete_list.append(k)
    delete_list.reverse()
    for j in delete_list:
        globals()[f'List{i}'].pop(j)
    globals()[f'List{i}'] = [remove_prefix(item) for item in globals()[f'List{i}']]

def Packing(i):
    Extraction(i,[2,5])
    globals()[f'List{i}'] = duplication(globals()[f'List{i}'])

def Compare(List_a, List_b, i):
    cnt = len(List_b)
    priceIs = len(List_b[0])
    for k in range(cnt):
        name = List_b[k][0]
        for j in range(len(List_a)):
            if List_a[j][0] == name:
                if priceIs==3:
                    if List_a[j][1] != List_b[k][1]:
                        print('---------------------------------------------------------------')
                        print(f'{sheet_List[i]} 시트에서 {List_b[k][0]} 제품의 수량이 틀렸습니다')
                        print('---------------------------------------------------------------')
                    if List_a[j][2] != List_b[k][2]:
                        print('---------------------------------------------------------------')
                        print(f'{sheet_List[i]} 시트에서 {List_b[k][0]} 제품의 가격이 틀렸습니다')
                        print('---------------------------------------------------------------')
                else:
                    if List_a[j][1] != List_b[k][1]:
                        print('---------------------------------------------------------------')
                        print(f'{sheet_List[i]} 시트에서 {List_b[k][0]} 제품의 수량이 틀렸습니다')
                        print('---------------------------------------------------------------')
    

for i in range(num_sheets):  
    total_cnt+=1
    # print(f'{sheet[i]} 검사')
    if '단가' in sheet_List[i]:
        Price(i)
    elif 'PI' in sheet_List[i]:
        if 'SHIPPING MARK' == sheet_List[i]:
            continue
        Proforma(i)
    elif '발주서' in sheet_List[i]:
            Order(i)
    elif 'CI' in sheet_List[i].upper() or 'INVOICE' in sheet_List[i].upper():
        Commercial(i)
    elif 'PL' in sheet_List[i].upper() or 'PACKING' in sheet_List[i].upper():
        Packing(i)
    else:
        print('-----------------------------------------')
        print('--------시트 명을 다시 확인해주세요--------')
        print('-----------------------------------------')


# for i in range(num_sheets):
#     if globals()[f'List{i}'] :
#         print(f'--------------------------------------------------')
#         print(f'---------------{sheet_List[i]}--------------------')
#         print(f'--------------------------------------------------')
#         print(globals()[f'List{i}'])

for j in range(len(List0)):
    Total_name_List.append(List0[j][0])
    if List0[j][1] != List1[j][1]:
        print('---------------------------------------------------------------')
        print(f'PI시트에서 {List1[j][0]} 제품의 수량이 틀렸음')
        print('---------------------------------------------------------------')
    if List0[j][2] != List1[j][2]:
        print('---------------------------------------------------------------')
        print(f'PI시트에서 {List1[j][0]} 제품의 가격이 틀렸음')
        print('---------------------------------------------------------------')

print('PI 정상적으로 작성되었습니다.')

for i in range(num_sheets-2):
    Compare(List0, globals()[f'List{i+2}'],i+2 )

print('모든 파일의 검사가 끝났습니다.')
